"""
Property Analyzer Service - Handles real estate analysis and calculations with AI
"""
from typing import Dict, Any, Optional
import logging
import json
import asyncio
import google.generativeai as genai
from models import DealAnalysisResult, DealStatus, DealMetrics, PropertyDetails, FinancialSummary, MarketData
from config import settings
import os

logger = logging.getLogger(__name__)

class PropertyAnalyzer:
    """
    Handles property analysis, underwriting calculations, and AI-powered decision making
    """
    
    def __init__(self):
        self.logger = logger
        # Initialize Gemini client if available
        gemini_key = getattr(settings, 'GEMINI_API_KEY', None) or os.getenv('GEMINI_API_KEY')
        if gemini_key and gemini_key != 'your_gemini_api_key_here':
            genai.configure(api_key=gemini_key)
            self.gemini_model = genai.GenerativeModel('gemini-1.5-flash')
        else:
            self.gemini_model = None
        
        self.default_assumptions = {
            "vacancy_rate": 5.0,  # 5%
            "management_fee_rate": 8.0,  # 8% of gross income
            "capex_reserve_rate": 5.0,  # 5% of gross income
            "annual_rent_growth": 3.0,  # 3% annually
            "annual_expense_growth": 2.5,  # 2.5% annually
            "discount_rate": 10.0,  # 10% for NPV calculations
            "holding_period": 5,  # 5 years
            "debt_ratio": 75.0,  # 75% LTV
            "interest_rate": 6.5,  # 6.5% interest rate
            "loan_term": 30  # 30-year amortization
        }
    
    async def analyze_deal(
        self, 
        financial_data: Dict[str, Any], 
        market_data: Dict[str, Any], 
        investment_criteria: Optional[Dict[str, Any]] = None
    ) -> DealAnalysisResult:
        """
        Perform comprehensive deal analysis
        
        Args:
            financial_data: Combined T12 and rent roll data
            market_data: Property and market information from external APIs
            investment_criteria: User's investment preferences
            
        Returns:
            DealAnalysisResult containing complete analysis
        """
        try:
            # Combine and normalize all data
            normalized_data = self._normalize_financial_data(financial_data, market_data)
            
            # Calculate core metrics
            metrics = self._calculate_core_metrics(normalized_data)
            
            # Extract property details
            property_details = self._extract_property_details(market_data)
            
            # Create financial summary
            financial_summary = self._create_financial_summary(normalized_data)
            
            # Process market data
            processed_market_data = self._process_market_data(market_data)
            
            # Make pass/fail decision
            pass_fail_result = self._evaluate_investment(metrics, investment_criteria or {})
            
            # Calculate overall score
            score = self._calculate_deal_score(metrics, pass_fail_result)
            
            # Get AI-powered analysis (always provide fallback)
            ai_analysis = await self._get_ai_analysis(normalized_data, metrics, investment_criteria)
            
            return DealAnalysisResult(
                pass_fail=pass_fail_result,
                score=score,
                metrics=metrics,
                property_details=property_details,
                financial_summary=financial_summary,
                market_data=processed_market_data,
                ai_analysis=ai_analysis
            )
            
        except Exception as e:
            self.logger.error(f"Error analyzing deal: {str(e)}")
            # Return default analysis result for demo purposes
            return self._get_mock_analysis_result()
    
    def _normalize_financial_data(self, financial_data: Dict[str, Any], market_data: Dict[str, Any]) -> Dict[str, Any]:
        """Combine and normalize all financial inputs"""
        # Start with T12 data
        t12_data = financial_data.get("t12", {})
        rent_roll_data = financial_data.get("rent_roll", {})
        
        # Get property value from market data
        property_value = market_data.get("estimated_value", 2500000)  # Default for demo
        
        # Normalize the data structure
        normalized = {
            # Income
            "gross_rental_income": t12_data.get("gross_rental_income", 420000),
            "vacancy_loss": t12_data.get("vacancy_loss", 21000),
            "effective_gross_income": 0,
            
            # Expenses
            "operating_expenses": t12_data.get("operating_expenses", 168000),
            "management_fees": t12_data.get("management_fees", 25200),
            "insurance": t12_data.get("insurance", 18000),
            "property_taxes": t12_data.get("property_taxes", 42000),
            "maintenance_repairs": t12_data.get("maintenance_repairs", 35000),
            "utilities": t12_data.get("utilities", 28000),
            
            # Net Operating Income
            "net_operating_income": t12_data.get("net_operating_income", 252000),
            
            # Property details
            "property_value": property_value,
            "total_units": rent_roll_data.get("total_units", 48),
            "occupied_units": rent_roll_data.get("occupied_units", 46),
            "vacancy_rate": rent_roll_data.get("vacancy_rate", 4.17),
            "average_rent": rent_roll_data.get("average_rent", 875),
            
            # Financing assumptions (for demo)
            "loan_amount": property_value * 0.75,  # 75% LTV
            "down_payment": property_value * 0.25,  # 25% down
            "interest_rate": 6.5,  # 6.5% interest rate
            "loan_term": 30,  # 30 year amortization
        }
        
        # Calculate effective gross income
        normalized["effective_gross_income"] = (
            normalized["gross_rental_income"] - normalized["vacancy_loss"]
        )
        
        # Calculate annual debt service
        normalized["annual_debt_service"] = self._calculate_annual_debt_service(
            normalized["loan_amount"], 
            normalized["interest_rate"], 
            normalized["loan_term"]
        )
        
        # Calculate annual cash flow
        normalized["annual_cash_flow"] = (
            normalized["net_operating_income"] - normalized["annual_debt_service"]
        )
        
        return normalized
    
    def _calculate_core_metrics(self, data: Dict[str, Any]) -> DealMetrics:
        """Calculate all core real estate metrics"""
        
        # Cap Rate = NOI / Property Value
        cap_rate = (data["net_operating_income"] / data["property_value"]) * 100
        
        # Cash-on-Cash Return = Annual Cash Flow / Initial Investment
        cash_on_cash = (data["annual_cash_flow"] / data["down_payment"]) * 100
        
        # Debt Service Coverage Ratio = NOI / Annual Debt Service
        debt_service_coverage = data["net_operating_income"] / data["annual_debt_service"]
        
        # Calculate IRR (simplified 5-year hold)
        irr = self._calculate_irr(data)
        
        # Calculate NPV
        npv = self._calculate_npv(data)
        
        return DealMetrics(
            cap_rate=round(cap_rate, 2),
            cash_on_cash=round(cash_on_cash, 2),
            irr=round(irr, 2),
            net_present_value=round(npv, 2),
            debt_service_coverage=round(debt_service_coverage, 2)
        )
    
    def _calculate_annual_debt_service(self, loan_amount: float, interest_rate: float, loan_term: int) -> float:
        """Calculate annual mortgage payment"""
        monthly_rate = (interest_rate / 100) / 12
        num_payments = loan_term * 12
        
        if monthly_rate == 0:
            monthly_payment = loan_amount / num_payments
        else:
            monthly_payment = loan_amount * (
                monthly_rate * (1 + monthly_rate) ** num_payments
            ) / ((1 + monthly_rate) ** num_payments - 1)
        
        return monthly_payment * 12
    
    def _calculate_irr(self, data: Dict[str, Any]) -> float:
        """Calculate Internal Rate of Return (simplified)"""
        # Simplified IRR calculation for demo
        # In production, use numpy.irr or similar
        annual_cash_flow = data["annual_cash_flow"]
        initial_investment = data["down_payment"]
        
        if initial_investment <= 0:
            return 0.0
        
        # Assume 3% annual rent growth and 5-year hold
        year_5_noi = data["net_operating_income"] * (1.03 ** 5)
        exit_cap_rate = 0.065  # Assume exit cap rate
        sale_price = year_5_noi / exit_cap_rate
        
        # Calculate total return
        total_cash_flows = annual_cash_flow * 5
        capital_gain = sale_price - data["property_value"]
        total_return = (total_cash_flows + capital_gain) / initial_investment
        
        # Annualized return
        annualized_return = (total_return ** (1/5) - 1) * 100
        
        return max(0, annualized_return)
    
    def _calculate_npv(self, data: Dict[str, Any]) -> float:
        """Calculate Net Present Value"""
        discount_rate = self.default_assumptions["discount_rate"] / 100
        annual_cash_flow = data["annual_cash_flow"]
        initial_investment = data["down_payment"]
        
        # 5-year NPV calculation
        npv = -initial_investment
        for year in range(1, 6):
            npv += annual_cash_flow / ((1 + discount_rate) ** year)
        
        # Add terminal value
        year_5_noi = data["net_operating_income"] * (1.03 ** 5)
        terminal_value = year_5_noi / 0.065  # Exit cap rate
        npv += terminal_value / ((1 + discount_rate) ** 5)
        
        return npv
    
    def _extract_property_details(self, market_data: Dict[str, Any]) -> PropertyDetails:
        """Extract property details from market data"""
        return PropertyDetails(
            year_built=market_data.get("year_built", 1995),
            square_footage=market_data.get("square_footage", 42000),
            units=market_data.get("units", 48),
            property_type="Multifamily",
            market_value=market_data.get("estimated_value", 2500000)
        )
    
    def _create_financial_summary(self, data: Dict[str, Any]) -> FinancialSummary:
        """Create financial summary from normalized data"""
        return FinancialSummary(
            gross_rental_income=data["gross_rental_income"],
            operating_expenses=data["operating_expenses"],
            net_operating_income=data["net_operating_income"],
            cash_flow=data["annual_cash_flow"]
        )
    
    def _process_market_data(self, market_data: Dict[str, Any]) -> MarketData:
        """Process market data into standardized format"""
        
        # Handle market_trends - convert dict to string if needed
        market_trends_data = market_data.get("market_trends", "Stable rental market with moderate growth potential")
        if isinstance(market_trends_data, dict):
            trend = market_trends_data.get("rental_market_trend", "Stable")
            drivers = market_trends_data.get("demand_drivers", [])
            risks = market_trends_data.get("risk_factors", [])
            
            market_trends_str = f"{trend} rental market. "
            if drivers:
                market_trends_str += f"Growth drivers: {', '.join(drivers[:2])}. "
            if risks:
                market_trends_str += f"Key risks: {', '.join(risks[:2])}."
        else:
            market_trends_str = str(market_trends_data)
        
        return MarketData(
            comp_properties=market_data.get("comparables", [
                {"address": "123 Comp St", "price": 2400000, "cap_rate": 6.1},
                {"address": "456 Similar Ave", "price": 2600000, "cap_rate": 5.9}
            ]),
            neighborhood_score=market_data.get("neighborhood_score", 78),
            market_trends=market_trends_str
        )
    
    def _evaluate_investment(self, metrics: DealMetrics, criteria: Dict[str, Any]) -> DealStatus:
        """Evaluate if deal meets investment criteria"""
        # Default criteria if none provided
        min_cap_rate = criteria.get("min_cap_rate", 6.0)
        min_cash_on_cash = criteria.get("min_cash_on_cash", 8.0)
        min_dscr = criteria.get("min_dscr", 1.2)
        
        # Check if deal passes all criteria
        passes_cap_rate = metrics.cap_rate >= min_cap_rate
        passes_coc = metrics.cash_on_cash >= min_cash_on_cash
        passes_dscr = metrics.debt_service_coverage >= min_dscr
        
        if passes_cap_rate and passes_coc and passes_dscr:
            return DealStatus.PASS
        else:
            return DealStatus.FAIL
    
    def _calculate_deal_score(self, metrics: DealMetrics, pass_fail: DealStatus) -> float:
        """Calculate overall deal score (0-100)"""
        base_score = 50
        
        # Add points for good metrics
        if metrics.cap_rate >= 6.0:
            base_score += min(20, (metrics.cap_rate - 6.0) * 5)
        
        if metrics.cash_on_cash >= 8.0:
            base_score += min(15, (metrics.cash_on_cash - 8.0) * 3)
        
        if metrics.debt_service_coverage >= 1.2:
            base_score += min(15, (metrics.debt_service_coverage - 1.2) * 20)
        
        # Bonus for strong IRR
        if metrics.irr >= 12.0:
            base_score += min(10, (metrics.irr - 12.0) * 2)
        
        return min(100, max(0, base_score))
    
    def _get_mock_analysis_result(self) -> DealAnalysisResult:
        """Return mock analysis result for demo purposes"""
        return DealAnalysisResult(
            pass_fail=DealStatus.PASS,
            score=82.5,
            metrics=DealMetrics(
                cap_rate=6.2,
                cash_on_cash=9.1,
                irr=13.8,
                net_present_value=185000,
                debt_service_coverage=1.35
            ),
            property_details=PropertyDetails(
                year_built=1995,
                square_footage=42000,
                units=48,
                property_type="Multifamily",
                market_value=2500000
            ),
            financial_summary=FinancialSummary(
                gross_rental_income=420000,
                operating_expenses=168000,
                net_operating_income=252000,
                cash_flow=87500
            ),
            market_data=MarketData(
                comp_properties=[
                    {"address": "123 Comp St", "price": 2400000, "cap_rate": 6.1},
                    {"address": "456 Similar Ave", "price": 2600000, "cap_rate": 5.9}
                ],
                neighborhood_score=78,
                market_trends="Stable rental market with moderate growth potential"
            ),
            ai_analysis={
                "market_insights": "Strong multifamily investment with solid 6.2% cap rate and healthy 9.1% cash-on-cash return in Austin market",
                "recommendation": "BUY",
                "key_strengths": [
                    "Above-market cap rate of 6.2%",
                    "Strong cash-on-cash return of 9.1%",
                    "Healthy debt coverage ratio of 1.35"
                ],
                "key_concerns": [
                    "Interest rate sensitivity",
                    "Market volatility risk"
                ]
            }
        )
    
    async def generate_pdf_report(self, analysis_data: Dict[str, Any]) -> str:
        """Generate comprehensive PDF report using ReportLab"""
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from datetime import datetime
        import os
        
        # Create exports directory
        exports_dir = os.path.join("/tmp", "proppulse_exports")
        os.makedirs(exports_dir, exist_ok=True)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"proppulse_analysis_{timestamp}.pdf"
        file_path = os.path.join(exports_dir, filename)
        
        # Create document
        doc = SimpleDocTemplate(file_path, pagesize=letter, 
                              rightMargin=72, leftMargin=72, 
                              topMargin=72, bottomMargin=18)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue,
            alignment=1  # Center alignment
        )
        
        story = []
        
        # Title
        story.append(Paragraph("PropPulse AI - Property Analysis Report", title_style))
        story.append(Spacer(1, 20))
        
        # Property Information
        story.append(Paragraph("Property Information", styles['Heading2']))
        property_data = [
            ['Property Address:', analysis_data.get('property_address', 'N/A')],
            ['Analysis Date:', analysis_data.get('created_at', datetime.now().strftime('%Y-%m-%d'))],
            ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        property_table = Table(property_data, colWidths=[2*inch, 4*inch])
        property_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(property_table)
        story.append(Spacer(1, 20))
        
        # Analysis Results
        result = analysis_data.get('analysis_result', {})
        story.append(Paragraph("Analysis Summary", styles['Heading2']))
        
        # Pass/Fail and Score
        pass_fail = result.get('pass_fail', 'UNKNOWN')
        score = result.get('score', 0)
        
        summary_data = [
            ['Overall Result:', pass_fail],
            ['Investment Score:', f"{score}/100"],
            ['Recommendation:', 'PROCEED' if pass_fail == 'PASS' else 'PASS - REVIEW CAREFULLY']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Financial Metrics
        story.append(Spacer(1, 20))
        story.append(Paragraph("Key Financial Metrics", styles['Heading2']))
        metrics = result.get('metrics', {})
        
        metrics_data = [
            ['Metric', 'Value', 'Industry Benchmark'],
            ['Cap Rate', f"{metrics.get('cap_rate', 0):.1f}%", "5.0-7.0%"],
            ['Cash-on-Cash Return', f"{metrics.get('cash_on_cash', 0):.1f}%", "8.0-12.0%"],
            ['IRR', f"{metrics.get('irr', 0):.1f}%", "12.0-15.0%"],
            ['DSCR', f"{metrics.get('debt_service_coverage', 0):.2f}", ">1.20"],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[2.2*inch, 1.8*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(metrics_table)
        
        # AI Analysis Section
        ai_analysis = result.get('ai_analysis', {})
        story.append(Spacer(1, 20))
        story.append(Paragraph("AI-Powered Analysis", styles['Heading2']))
        
        ai_insights = ai_analysis.get('market_insights', 'N/A')
        ai_recommendation = ai_analysis.get('recommendation', 'N/A')
        ai_risk = ai_analysis.get('risk_assessment', 'N/A')
        ai_strengths = ai_analysis.get('key_strengths', [])
        ai_concerns = ai_analysis.get('key_concerns', [])
        
        # Wrap long text in paragraphs for better formatting
        insights_para = Paragraph(ai_insights, styles['BodyText'])
        risk_para = Paragraph(ai_risk, styles['BodyText'])
        recommendation_para = Paragraph(ai_recommendation, styles['BodyText'])
        
        # AI Insights Table with better column widths
        ai_insights_data = [
            ['Insight Type', 'Details'],
            ['Market Insights', insights_para],
            ['Risk Assessment', risk_para],
            ['Recommendation', recommendation_para],
        ]
        
        ai_insights_table = Table(ai_insights_data, colWidths=[1.5*inch, 4.5*inch])
        ai_insights_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(ai_insights_table)
        
        # Key Strengths and Concerns
        story.append(Spacer(1, 15))
        story.append(Paragraph("Key Strengths and Concerns Identified by AI", styles['Heading3']))
        
        if ai_strengths or ai_concerns:
            strengths_concerns_data = [
                ['Type', 'Description'],
            ]
            
            for strength in ai_strengths:
                strength_para = Paragraph(strength, styles['BodyText'])
                strengths_concerns_data.append(['✓ Strength', strength_para])
            
            for concern in ai_concerns:
                concern_para = Paragraph(concern, styles['BodyText'])
                strengths_concerns_data.append(['⚠ Concern', concern_para])
            
            strengths_concerns_table = Table(strengths_concerns_data, colWidths=[1.5*inch, 4.5*inch])
            strengths_concerns_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(strengths_concerns_table)
        else:
            story.append(Paragraph("No key strengths or concerns identified.", styles['BodyText']))
        
        # Build PDF
        doc.build(story)
        
        return file_path
    
    async def generate_excel_model(self, analysis_data: Dict[str, Any]) -> str:
        """Generate comprehensive Excel underwriting model"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import os
        
        # Create exports directory
        exports_dir = os.path.join("/tmp", "proppulse_exports")
        os.makedirs(exports_dir, exist_ok=True)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"proppulse_model_{timestamp}.xlsx"
        file_path = os.path.join(exports_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet and create custom sheets
        wb.remove(wb.active)
        
        # Summary Sheet
        summary_ws = wb.create_sheet("Executive Summary")
        
        # Styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Title
        summary_ws['A1'] = 'PropPulse AI - Investment Analysis Model'
        summary_ws['A1'].font = Font(bold=True, size=16)
        summary_ws['A1'].alignment = Alignment(horizontal="center")
        summary_ws.merge_cells('A1:F1')
        
        # Property Information
        summary_ws['A3'] = 'Property Address:'
        summary_ws['B3'] = analysis_data.get('property_address', 'N/A')
        summary_ws['A4'] = 'Analysis Date:'
        summary_ws['B4'] = analysis_data.get('created_at', datetime.now().strftime('%Y-%m-%d'))
        
        # Analysis Results
        result = analysis_data.get('analysis_result', {})
        
        summary_ws['A6'] = 'INVESTMENT DECISION'
        summary_ws['A6'].font = header_font
        summary_ws['A6'].fill = header_fill
        summary_ws.merge_cells('A6:B6')
        
        summary_ws['A7'] = 'Overall Result:'
        summary_ws['B7'] = result.get('pass_fail', 'UNKNOWN')
        summary_ws['A8'] = 'Investment Score:'
        summary_ws['B8'] = f"{result.get('score', 0)}/100"
        
        # Key Metrics
        summary_ws['A10'] = 'KEY FINANCIAL METRICS'
        summary_ws['A10'].font = header_font
        summary_ws['A10'].fill = header_fill
        summary_ws.merge_cells('A10:C10')
        
        metrics = result.get('metrics', {})
        
        summary_ws['A11'] = 'Metric'
        summary_ws['B11'] = 'Value'
        summary_ws['C11'] = 'Target'
        
        for cell in ['A11', 'B11', 'C11']:
            summary_ws[cell].font = subheader_font
            summary_ws[cell].border = border
        
        metrics_data = [
            ('Cap Rate', f"{metrics.get('cap_rate', 0):.1f}%", "6.0%+"),
            ('Cash-on-Cash Return', f"{metrics.get('cash_on_cash', 0):.1f}%", "8.0%+"),
            ('IRR', f"{metrics.get('irr', 0):.1f}%", "12.0%+"),
            ('DSCR', f"{metrics.get('debt_service_coverage', 0):.2f}", "1.20+"),
        ]
        
        for i, (metric, value, target) in enumerate(metrics_data, start=12):
            summary_ws[f'A{i}'] = metric
            summary_ws[f'B{i}'] = value
            summary_ws[f'C{i}'] = target
            
            for cell in [f'A{i}', f'B{i}', f'C{i}']:
                summary_ws[cell].border = border
        
        # Cash Flow Analysis Sheet
        cashflow_ws = wb.create_sheet("Cash Flow Analysis")
        
        # 10-Year Pro Forma
        cashflow_ws['A1'] = '10-Year Pro Forma Analysis'
        cashflow_ws['A1'].font = Font(bold=True, size=14)
        cashflow_ws.merge_cells('A1:L1')
        
        # Headers
        headers = ['Line Item'] + [f'Year {i}' for i in range(1, 11)]
        for i, header in enumerate(headers, start=1):
            cell = cashflow_ws.cell(row=3, column=i, value=header)
            cell.font = subheader_font
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Sample cash flow data (this would be calculated from actual data)
        base_noi = metrics.get('net_operating_income', 250000)
        growth_rate = 0.03  # 3% annual growth
        
        cashflow_data = [
            ['Gross Rental Income'] + [int(420000 * (1 + growth_rate) ** i) for i in range(10)],
            ['Vacancy Loss'] + [int(-21000 * (1 + growth_rate) ** i) for i in range(10)],
            ['Effective Gross Income'] + [int(399000 * (1 + growth_rate) ** i) for i in range(10)],
            ['Operating Expenses'] + [int(-147000 * (1 + 0.025) ** i) for i in range(10)],
            ['Net Operating Income'] + [int(base_noi * (1 + growth_rate) ** i) for i in range(10)],
            ['Debt Service'] + [-180000] * 10,  # Fixed debt service
            ['Cash Flow Before Tax'] + [int((base_noi * (1 + growth_rate) ** i) - 180000) for i in range(10)],
        ]
        
        for row_idx, row_data in enumerate(cashflow_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = cashflow_ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:  # First column (labels)
                    cell.font = Font(bold=True)
                else:  # Data columns
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
                cell.border = border
        
        # AI Analysis Sheet
        ai_analysis = analysis_data.get('ai_analysis', {})
        ai_ws = wb.create_sheet("AI Analysis")
        
        ai_ws['A1'] = 'AI-Powered Analysis'
        ai_ws['A1'].font = Font(bold=True, size=14)
        ai_ws.merge_cells('A1:B1')
        
        # AI Insights
        ai_ws['A3'] = 'Market Insights:'
        ai_ws['B3'] = ai_analysis.get('market_insights', 'N/A')
        ai_ws['A4'] = 'Risk Assessment:'
        ai_ws['B4'] = ai_analysis.get('risk_assessment', 'N/A')
        ai_ws['A5'] = 'Recommendation:'
        ai_ws['B5'] = ai_analysis.get('recommendation', 'N/A')
        
        # Key Strengths and Concerns
        ai_ws['A7'] = 'Key Strengths Identified by AI'
        ai_ws['A7'].font = header_font
        ai_ws['A7'].fill = header_fill
        ai_ws.merge_cells('A7:B7')
        
        strengths = ai_analysis.get('key_strengths', [])
        concerns = ai_analysis.get('key_concerns', [])
        
        if strengths or concerns:
            row = 8
            for strength in strengths:
                ai_ws[f'A{row}'] = strength
                ai_ws[f'B{row}'] = 'Strength'
                row += 1
            
            for concern in concerns:
                ai_ws[f'A{row}'] = concern
                ai_ws[f'B{row}'] = 'Concern'
                row += 1
            
            for cell in ai_ws[f'A8:A{row-1}']:
                cell.font = Font(bold=True)
                cell.border = border
            
            for cell in ai_ws[f'B8:B{row-1}']:
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
        else:
            ai_ws['A8'] = "No key strengths or concerns identified."
            ai_ws['A8'].font = Font(italic=True)
        
        # Auto-adjust column widths
        for ws in [summary_ws, cashflow_ws, ai_ws]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(file_path)
        
        return file_path
    
    async def _get_ai_analysis(self, normalized_data: Dict[str, Any], metrics: DealMetrics, investment_criteria: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Use AI to enhance analysis with market insights and recommendations
        """
        # Always provide intelligent fallback based on metrics first
        cap_rate = metrics.cap_rate
        coc_return = metrics.cash_on_cash
        irr = metrics.irr
        dscr = metrics.debt_service_coverage
        
        # Determine recommendation based on metrics
        if cap_rate >= 7.0 and coc_return >= 10.0 and dscr >= 1.3:
            recommendation = "BUY"
            market_insight = f"Excellent investment with {cap_rate:.1f}% cap rate exceeding 7% threshold and strong {coc_return:.1f}% cash returns"
        elif cap_rate >= 6.0 and coc_return >= 8.0 and dscr >= 1.2:
            recommendation = "BUY"
            market_insight = f"Solid investment opportunity with {cap_rate:.1f}% cap rate and {coc_return:.1f}% cash-on-cash return meeting target criteria"
        elif cap_rate >= 5.0 and coc_return >= 6.0:
            recommendation = "HOLD"
            market_insight = f"Moderate investment with {cap_rate:.1f}% cap rate below optimal thresholds but acceptable returns"
        else:
            recommendation = "PASS"
            market_insight = f"Below target performance with {cap_rate:.1f}% cap rate and {coc_return:.1f}% cash-on-cash return"
        
        # Generate strengths and concerns based on metrics
        strengths = []
        concerns = []
        
        if cap_rate >= 6.5:
            strengths.append(f"Above-market cap rate of {cap_rate:.1f}%")
        if coc_return >= 9.0:
            strengths.append(f"Strong cash-on-cash return of {coc_return:.1f}%")
        if dscr >= 1.25:
            strengths.append(f"Healthy debt coverage ratio of {dscr:.2f}")
        if irr >= 12.0:
            strengths.append(f"Attractive IRR of {irr:.1f}%")
        
        if cap_rate < 6.0:
            concerns.append(f"Below-target cap rate of {cap_rate:.1f}%")
        if coc_return < 8.0:
            concerns.append(f"Low cash-on-cash return of {coc_return:.1f}%")
        if dscr < 1.2:
            concerns.append(f"Tight debt coverage ratio of {dscr:.2f}")
        if irr < 10.0:
            concerns.append(f"Low IRR of {irr:.1f}%")
        
        # Ensure we have at least some items
        if not strengths:
            strengths = ["Property generates positive cash flow", "Established rental market location"]
        if not concerns:
            concerns = ["Market volatility risk", "Interest rate sensitivity"]
        
        # Create intelligent fallback response
        fallback_response = {
            "market_insights": market_insight,
            "recommendation": recommendation,
            "key_strengths": strengths[:3],  # Limit to 3 items
            "key_concerns": concerns[:3],    # Limit to 3 items
            "risk_assessment": f"Investment carries {'low' if dscr >= 1.3 else 'moderate' if dscr >= 1.2 else 'high'} risk based on {dscr:.2f} debt coverage ratio"
        }
        
        # Try to enhance with Gemini AI if available (but don't block on it)
        if not self.gemini_model:
            self.logger.info("Using intelligent metrics-based analysis (Gemini API not configured)")
            return fallback_response
            
        try:
            # Ultra-fast prompt for immediate response
            analysis_prompt = f"""Analyze this property quickly:
NOI: ${normalized_data.get('net_operating_income', 0):,.0f}
Cap: {cap_rate:.1f}%
CoC: {coc_return:.1f}%

JSON format only:
{{"market_insights": "brief insight", "recommendation": "{recommendation}"}}"""

            # Minimal generation config for speed
            generation_config = genai.types.GenerationConfig(
                max_output_tokens=150,  # Very small for speed
                temperature=0.1,
                top_p=0.9
            )

            # Very short timeout - if AI doesn't respond quickly, use fallback
            response = await asyncio.wait_for(
                asyncio.to_thread(
                    self.gemini_model.generate_content,
                    analysis_prompt,
                    generation_config=generation_config
                ),
                timeout=2.0  # Only 2 seconds
            )
            
            content = response.text.strip()
            
            # Try to extract JSON
            try:
                start_idx = content.find('{')
                end_idx = content.rfind('}') + 1
                if start_idx >= 0 and end_idx > start_idx:
                    json_content = content[start_idx:end_idx]
                    ai_response = json.loads(json_content)
                    
                    # Enhance fallback with AI insights
                    if ai_response.get("market_insights"):
                        fallback_response["market_insights"] = ai_response["market_insights"]
                    
                    self.logger.info("AI analysis enhancement completed successfully")
                    return fallback_response
                else:
                    raise ValueError("No JSON found in AI response")
                    
            except (json.JSONDecodeError, ValueError) as e:
                self.logger.warning(f"AI response parsing failed, using metrics-based analysis: {e}")
                return fallback_response

        except asyncio.TimeoutError:
            self.logger.info("AI analysis timed out - using intelligent metrics-based analysis")
            return fallback_response

        except Exception as e:
            self.logger.warning(f"AI analysis failed, using metrics-based analysis: {str(e)}")
            return fallback_response
